import streamlit as st
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook

st.set_page_config(page_title="GSTR-1 Processor", layout="centered")
st.title("GSTR-1 Excel Processor")

# ---------------- Utilities ---------------- #

def normalize_columns(df):
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    return df


def find_column_by_keywords(df, keywords, label):
    for col in df.columns:
        col_l = col.lower()
        if all(k.lower() in col_l for k in keywords):
            return col
    raise KeyError(
        f"{label} column not found.\n"
        f"Expected keywords: {keywords}\n"
        f"Found columns: {list(df.columns)}"
    )


def get_column_letter_by_header(ws, header_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return ws.cell(row=1, column=col).column_letter
    raise KeyError(f"Column '{header_name}' not found in sheet '{ws.title}'")

# ---------------- Session State ---------------- #

if "processed" not in st.session_state:
    st.session_state.processed = False
    st.session_state.outputs = {}

# ---------------- UI ---------------- #

company_code = st.text_input("Company Code (XXXX)")

sd_file = st.file_uploader("Upload SD File", type="xlsx")
sr_file = st.file_uploader("Upload SR File", type="xlsx")
tb_file = st.file_uploader("Upload TB File", type="xlsx")
gl_file = st.file_uploader("Upload GL Dump File", type="xlsx")

# ---------------- Processing ---------------- #

if st.button("Process Files"):

    if not company_code:
        st.error("Company Code is mandatory")
        st.stop()

    if not all([sd_file, sr_file, tb_file, gl_file]):
        st.error("All files must be uploaded")
        st.stop()

    try:
        outputs = {}

        with tempfile.TemporaryDirectory() as tmpdir:

            # Save uploaded files
            paths = {}
            for f, name in [
                (sd_file, "sd.xlsx"),
                (sr_file, "sr.xlsx"),
                (tb_file, "tb.xlsx"),
                (gl_file, "gl.xlsx"),
            ]:
                path = os.path.join(tmpdir, name)
                with open(path, "wb") as out:
                    out.write(f.getbuffer())
                paths[name] = path

            # ---------- SALES REGISTER ---------- #

            df_sd = normalize_columns(pd.read_excel(paths["sd.xlsx"]))
            df_sr = normalize_columns(pd.read_excel(paths["sr.xlsx"]))
            df_sales = pd.concat([df_sd, df_sr], ignore_index=True)

            # ---------- GL ---------- #

            df_gl = normalize_columns(pd.read_excel(paths["gl.xlsx"]))

            gl_text_col = find_column_by_keywords(df_gl, ["g/l", "account", "long", "text"], "GL Text")
            gl_account_col = find_column_by_keywords(df_gl, ["g/l", "account"], "GL Account")
            value_col = find_column_by_keywords(df_gl, ["value"], "Amount")
            doc_col = find_column_by_keywords(df_gl, ["document"], "Document Number")

            gst_accounts = [
                "Central GST Payable",
                "Integrated GST Payable",
                "State GST Payable",
            ]

            df_gst = df_gl[df_gl[gl_text_col].isin(gst_accounts)]
            df_revenue = df_gl[df_gl[gl_account_col].astype(str).str.startswith("3")]

            # ---------- TB ---------- #

            df_tb = normalize_columns(pd.read_excel(paths["tb.xlsx"]))

            tb_text_col = find_column_by_keywords(df_tb, ["g/l", "acct", "long", "text"], "TB GL Text")
            debit_col = find_column_by_keywords(df_tb, ["period", "d"], "TB Debit")
            credit_col = find_column_by_keywords(df_tb, ["period", "c"], "TB Credit")

            df_tb_gst = df_tb[df_tb[tb_text_col].isin(gst_accounts)].copy()
            df_tb_gst["Difference as per TB"] = df_tb_gst[credit_col] - df_tb_gst[debit_col]

            # ---------- GST SUMMARY ---------- #

            gst_summary_df = (
                df_gst
                .groupby(gl_text_col, as_index=False)[value_col]
                .sum()
                .rename(columns={
                    gl_text_col: "GST Type",
                    value_col: "GST Payable as per GL"
                })
            )

            tb_summary_df = (
                df_tb_gst
                .groupby(tb_text_col, as_index=False)["Difference as per TB"]
                .sum()
                .rename(columns={tb_text_col: "GST Type"})
            )

            summary_df = pd.merge(
                gst_summary_df,
                tb_summary_df,
                on="GST Type",
                how="left"
            ).fillna(0)

            # ---------- GSTR-1 WORKBOOK ---------- #

            gstr_path = os.path.join(tmpdir, f"{company_code}_GSTR-1_Workbook.xlsx")

            with pd.ExcelWriter(gstr_path, engine="openpyxl") as writer:
                df_sales.to_excel(writer, sheet_name="Sales register", index=False)
                df_revenue.to_excel(writer, sheet_name="Revenue", index=False)
                df_gst.to_excel(writer, sheet_name="GST payable", index=False)
                summary_df.to_excel(writer, sheet_name="GST Summary", index=False)

            # ---------- ADD FORMULAS & FORMATS ---------- #

            wb = load_workbook(gstr_path)

            ws_sales = wb["Sales register"]
            ws_rev = wb["Revenue"]
            ws_gst = wb["GST payable"]
            ws_summary = wb["GST Summary"]

            # --- Column detection for vlookup ---
            sales_lookup_col = get_column_letter_by_header(ws_sales, "Generic Field 8")
            rev_doc_col = get_column_letter_by_header(ws_rev, "Document Number")
            gst_doc_col = get_column_letter_by_header(ws_gst, "Document Number")

            # --- Sales register vlookups ---
            sales_last_col = ws_sales.max_column
            ws_sales.cell(1, sales_last_col + 1, "Revenue VLOOKUP")
            ws_sales.cell(1, sales_last_col + 2, "GST Payable VLOOKUP")

            for r in range(2, ws_sales.max_row + 1):
                ws_sales.cell(
                    r, sales_last_col + 1,
                    f'=IFERROR(VLOOKUP({sales_lookup_col}{r},Revenue!{rev_doc_col}:{rev_doc_col},1,FALSE),"Not Found")'
                )
                ws_sales.cell(
                    r, sales_last_col + 2,
                    f'=IFERROR(VLOOKUP({sales_lookup_col}{r},\'GST payable\'!{gst_doc_col}:{gst_doc_col},1,FALSE),"Not Found")'
                )

            # --- Cross vlookups ---
            rev_last_col = ws_rev.max_column
            ws_rev.cell(1, rev_last_col + 1, "Sales Register VLOOKUP")
            for r in range(2, ws_rev.max_row + 1):
                ws_rev.cell(
                    r, rev_last_col + 1,
                    f'=IFERROR(VLOOKUP({rev_doc_col}{r},\'Sales register\'!{sales_lookup_col}:{sales_lookup_col},1,FALSE),"Not Found")'
                )

            gst_last_col = ws_gst.max_column
            ws_gst.cell(1, gst_last_col + 1, "Sales Register VLOOKUP")
            for r in range(2, ws_gst.max_row + 1):
                ws_gst.cell(
                    r, gst_last_col + 1,
                    f'=IFERROR(VLOOKUP({gst_doc_col}{r},\'Sales register\'!{sales_lookup_col}:{sales_lookup_col},1,FALSE),"Not Found")'
                )

            # --- Net Difference formula + NUMBER format ---
            ws_summary.cell(1, 4, "Net Difference")
            for r in range(2, ws_summary.max_row + 1):
                cell = ws_summary.cell(r, 4, f"=B{r}+C{r}")
                cell.number_format = "0.00"

            wb.save(gstr_path)

            with open(gstr_path, "rb") as f:
                outputs["GSTR-1 Workbook.xlsx"] = f.read()

        st.session_state.outputs = outputs
        st.session_state.processed = True
        st.success("Processing completed successfully")

    except Exception as e:
        st.error(str(e))

# ---------------- Downloads ---------------- #

if st.session_state.processed:
    for filename, data in st.session_state.outputs.items():
        st.download_button(
            label=f"Download {filename}",
            data=data,
            file_name=filename,
            key=filename
        )







